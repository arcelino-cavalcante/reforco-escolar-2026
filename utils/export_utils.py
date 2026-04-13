"""
Utilitários de exportação de relatórios em PDF e Excel.
Usa fpdf2 para PDFs e openpyxl (via pandas) para Excel.
"""
import io
import datetime
import pandas as pd
from fpdf import FPDF
from database.crud import compreensao_label


# ==========================================
# CONFIGURAÇÕES VISUAIS DO PDF
# ==========================================
COR_HEADER = (44, 62, 80)       # #2C3E50
COR_HEADER_CLARO = (52, 152, 219)  # #3498DB
COR_LINHA_PAR = (245, 247, 250)
COR_LINHA_IMPAR = (255, 255, 255)
COR_VERDE = (46, 204, 113)
COR_VERMELHO = (231, 76, 60)


class RelatorioPDF(FPDF):
    """Classe base para todos os PDFs do sistema, com header/footer padronizados."""

    def __init__(self, titulo_relatorio="Relatório", subtitulo=""):
        super().__init__()
        self.titulo_relatorio = titulo_relatorio
        self.subtitulo = subtitulo
        self.set_auto_page_break(auto=True, margin=20)

    def header(self):
        self.set_font("Helvetica", "B", 16)
        self.set_text_color(*COR_HEADER)
        self.cell(0, 10, self.titulo_relatorio, ln=True, align="C")
        if self.subtitulo:
            self.set_font("Helvetica", "", 10)
            self.set_text_color(127, 140, 141)
            self.cell(0, 6, self.subtitulo, ln=True, align="C")
        self.set_draw_color(*COR_HEADER_CLARO)
        self.set_line_width(0.5)
        self.line(10, self.get_y() + 2, self.w - 10, self.get_y() + 2)
        self.ln(6)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(170, 170, 170)
        data_hora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
        self.cell(0, 10, f"Reforco Escolar - Gerado em {data_hora} | Pagina {self.page_no()}/{{nb}}", align="C")

    def secao(self, titulo):
        """Renderiza um título de seção formatado."""
        self.set_font("Helvetica", "B", 13)
        self.set_text_color(*COR_HEADER_CLARO)
        self.cell(0, 10, titulo, ln=True)
        self.set_draw_color(*COR_HEADER_CLARO)
        self.set_line_width(0.3)
        self.line(10, self.get_y(), self.w - 10, self.get_y())
        self.ln(3)

    def info_box(self, label, valor):
        """Renderiza um par label: valor."""
        self.set_font("Helvetica", "B", 10)
        self.set_text_color(*COR_HEADER)
        self.cell(55, 7, f"{label}:", align="R")
        self.set_font("Helvetica", "", 10)
        self.set_text_color(50, 50, 50)
        self.cell(0, 7, f"  {valor}", ln=True)

    def tabela(self, colunas, dados, larguras=None):
        """Renderiza uma tabela com cores alternadas."""
        if not larguras:
            largura_disponivel = self.w - 20
            larguras = [largura_disponivel / len(colunas)] * len(colunas)

        # Header da tabela
        self.set_fill_color(*COR_HEADER)
        self.set_text_color(255, 255, 255)
        self.set_font("Helvetica", "B", 9)
        for i, col in enumerate(colunas):
            self.cell(larguras[i], 8, col, border=1, fill=True, align="C")
        self.ln()

        # Linhas de dados
        self.set_font("Helvetica", "", 8)
        for idx, linha in enumerate(dados):
            cor = COR_LINHA_PAR if idx % 2 == 0 else COR_LINHA_IMPAR
            self.set_fill_color(*cor)
            self.set_text_color(50, 50, 50)

            for i, val in enumerate(linha):
                texto = str(val) if val is not None else ""
                # Truncar se muito longo
                max_chars = int(larguras[i] / 2.2)
                if len(texto) > max_chars:
                    texto = texto[:max_chars - 2] + ".."
                self.cell(larguras[i], 7, texto, border=1, fill=True, align="C")
            self.ln()


# ==========================================
# PDF: DOSSIÊ DO ESTUDANTE
# ==========================================
def gerar_pdf_dossie_estudante(nome_aluno, turma_nome, etapa_nome, historico_diario, historico_mensal):
    """Gera PDF completo do dossiê de um estudante (para reuniões de pais)."""
    pdf = RelatorioPDF(
        titulo_relatorio="Dossie do Estudante",
        subtitulo=f"{nome_aluno} - {turma_nome} ({etapa_nome})"
    )
    pdf.alias_nb_pages()
    pdf.add_page()

    # Dados gerais
    pdf.secao("Dados do Estudante")
    pdf.info_box("Nome", nome_aluno)
    pdf.info_box("Turma", turma_nome)
    pdf.info_box("Etapa", etapa_nome)
    data_hoje = datetime.date.today().strftime("%d/%m/%Y")
    pdf.info_box("Data do Relatorio", data_hoje)

    # Métricas resumo
    if historico_diario:
        total = len(historico_diario)
        presencas = sum(1 for r in historico_diario if r.get('compareceu') == 1)
        faltas = total - presencas
        taxa = (presencas / total * 100) if total > 0 else 0

        pdf.ln(3)
        pdf.secao("Resumo de Frequencia")
        pdf.info_box("Total de Lancamentos", str(total))
        pdf.info_box("Presencas", str(presencas))
        pdf.info_box("Faltas", str(faltas))
        pdf.info_box("Taxa de Presenca", f"{taxa:.1f}%")

    # Consolidados Bimestrais
    if historico_mensal:
        pdf.ln(3)
        pdf.secao("Fechamentos Bimestrais")
        for cons in historico_mensal:
            bim = cons.get('bimestre', '?')
            parecer = cons.get('parecer_evolutivo', 'Nao informado')
            alta = "SIM" if cons.get('recomendacao_alta') else "Nao"

            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(*COR_HEADER)
            pdf.cell(0, 7, f"Bimestre {bim}", ln=True)
            pdf.info_box("Parecer", parecer)
            pdf.info_box("Recomendacao de Alta", alta)

            if cons.get('acao_pedagogica'):
                pdf.info_box("Acao Pedagogica", cons.get('acao_pedagogica', ''))

            # Notas de Matemática
            if cons.get('mat_adicao') is not None:
                notas_mat = f"Adicao({cons.get('mat_adicao')}), Subtracao({cons.get('mat_subtracao')}), Multiplicacao({cons.get('mat_multiplicacao')}), Divisao({cons.get('mat_divisao')})"
                pdf.info_box("Matematica", notas_mat)

            # Notas de Português
            if cons.get('port_leitura') is not None:
                notas_port = f"Leitura({cons.get('port_leitura')}), Escrita({cons.get('port_escrita')}), Interpretacao({cons.get('port_interpretacao')})"
                pdf.info_box("Portugues", notas_port)

            pdf.ln(2)

    # Histórico Diário (Timeline)
    if historico_diario:
        pdf.ln(3)
        pdf.secao("Historico de Aulas no Reforco")

        colunas = ["Data", "Status", "Habilidade", "Compreensao", "Obs."]
        larguras = [25, 22, 50, 45, 48]
        dados = []

        for reg in historico_diario:
            data_f = datetime.datetime.strptime(reg['data_registro'], '%Y-%m-%d').strftime('%d/%m/%Y')
            status = "Presente" if reg['compareceu'] == 1 else "Ausente"

            if reg['compareceu'] == 1:
                hab = reg.get('habilidade_trabalhada', '')
                nv = compreensao_label(reg.get('nivel_compreensao', 0))
                obs = reg.get('observacao', '') or ''
            else:
                hab = f"Falta: {reg.get('motivo_falta', '')}"
                nv = "-"
                obs = ""

            dados.append([data_f, status, hab, nv, obs])

        pdf.tabela(colunas, dados, larguras)

    # Retornar bytes
    return bytes(pdf.output())


# ==========================================
# PDF: RELATÓRIO DO ALUNO (PAINEL REGENTE)
# ==========================================
def gerar_pdf_aluno_regente(nome_aluno, turma_nome, historico_aulas, consolidados):
    """Gera PDF do relatório de um aluno específico para o consel de classe (visão do regente)."""
    pdf = RelatorioPDF(
        titulo_relatorio="Relatorio para Conselho de Classe",
        subtitulo=f"Estudante: {nome_aluno} - Turma: {turma_nome}"
    )
    pdf.alias_nb_pages()
    pdf.add_page()

    # Resumo
    pdf.secao("Dados do Estudante")
    pdf.info_box("Nome", nome_aluno)
    pdf.info_box("Turma", turma_nome)

    if historico_aulas:
        total = len(historico_aulas)
        presencas = sum(1 for r in historico_aulas if r.get('compareceu') == 1)
        taxa = (presencas / total * 100) if total > 0 else 0
        pdf.info_box("Total de Atendimentos", str(total))
        pdf.info_box("Taxa de Presenca", f"{taxa:.1f}%")

    # Consolidados
    if consolidados:
        pdf.ln(3)
        pdf.secao("Parecer Bimestral (Equipe de Reforco)")
        for cons in consolidados:
            bim = cons.get('bimestre', '?')
            parecer = cons.get('parecer_evolutivo', 'Nao informado')
            prof_ref = cons.get('prof_reforco_nome', 'Equipe')
            alta = "SIM - ALUNO APTO PARA SALA REGULAR" if cons.get('recomendacao_alta') else "Nao"

            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(*COR_HEADER)
            pdf.cell(0, 7, f"Bimestre {bim} (Prof. Reforco: {prof_ref})", ln=True)
            pdf.info_box("Parecer Evolutivo", parecer)
            pdf.info_box("Recomendacao de Alta", alta)

            if cons.get('acao_pedagogica'):
                pdf.info_box("Acao Pedagogica Sugerida", cons.get('acao_pedagogica', ''))
            pdf.ln(2)

    # Timeline de aulas
    if historico_aulas:
        pdf.ln(3)
        pdf.secao("Diario de Aulas no Reforco")

        colunas = ["Data", "Bim", "Status", "Habilidade", "Compreensao", "Prof. Reforco"]
        larguras = [23, 12, 20, 45, 42, 48]
        dados = []

        for log in historico_aulas:
            data_f = datetime.datetime.strptime(log['data_registro'], '%Y-%m-%d').strftime('%d/%m/%Y')
            bim = log.get('bimestre', '')
            status = "Presente" if log['compareceu'] == 1 else "Ausente"
            prof = f"{log.get('prof_reforco_nome', '')} ({log.get('prof_reforco_area', '')})"

            if log['compareceu'] == 1:
                hab = log.get('habilidade_trabalhada', '')
                nv = compreensao_label(log.get('nivel_compreensao', 0))
            else:
                hab = log.get('motivo_falta', '')
                nv = "-"

            dados.append([data_f, bim, status, hab, nv, prof])

        pdf.tabela(colunas, dados, larguras)

    return bytes(pdf.output())


# ==========================================
# EXCEL: AUDITORIA DE REGISTROS
# ==========================================
def gerar_excel_auditoria(df_clean):
    """Gera Excel formatado a partir de um DataFrame da auditoria de registros."""
    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_clean.to_excel(writer, index=False, sheet_name='Registros')

        # Formatar a planilha
        ws = writer.sheets['Registros']

        # Ajustar larguras de coluna baseado no conteúdo
        for col_idx, col in enumerate(df_clean.columns, 1):
            max_length = max(
                df_clean[col].astype(str).map(len).max() if len(df_clean) > 0 else 0,
                len(col)
            )
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = min(max_length + 4, 50)

        # Estilizar header
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    buffer.seek(0)
    return buffer.getvalue()
